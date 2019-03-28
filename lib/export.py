import openpyxl
import os
from statistics import mean
from datetime import datetime
import json
import shutil
from zipfile import ZipFile

from conf import my_setting
from lib.common import list_split
from Sewage import settings


def ex_container(monitor_objs):
    # for m in monitor_objs:
    #     monitor_name = m.name
    #     monitor_people = m.people
    #     monitor_date = m.start_time
    #     sample_objs = m.sample.all()
    # print('container')
    date = datetime.now().date()
    output_path = os.path.join(my_setting.export_folder, str(date))
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    static_excel = os.path.join(output_path, '小区监测统计表.xlsx')
    for count_m, m in enumerate(monitor_objs, 1):  # 遍历所有监测点
        output_path_per_monitor = os.path.join(output_path, m.name)

        if not os.path.exists(output_path_per_monitor):
            os.mkdir(output_path_per_monitor)
        monitor_flow_dic = {}
        exterior_photo = json.loads(m.exterior_photo)
        exterior_photo = list_split(exterior_photo)
        for ep in exterior_photo:
            shutil.copy('{}{}{}'.format(settings.BASE_DIR, os.sep, ep), output_path_per_monitor)

        fs = m.flow.all().distinct('flow_date')  # 根据采样时间分组
        f_date_lst = []
        # export_excel_path = os.path.join()
        if count_m == 1:
            m_wb = openpyxl.load_workbook(os.path.join(my_setting.excel_folder, '小区监测统计表.xlsx'))
            f_ws = m_wb['流量']
            s_ws = m_wb['水质']
        else:
            m_wb = openpyxl.load_workbook(static_excel)
            f_ws = m_wb['流量']
            s_ws = m_wb['水质']
        for f in fs:
            f_date_lst.append(f.flow_date)  # 所有的采样日期
        for i in f_date_lst:
            if s_ws['H{}'.format(4 + 3 * count_m - 2)].value != '' and s_ws[
                'H{}'.format(4 + 3 * count_m - 2)].value is not None and s_ws[
                'H{}'.format(4 + 3 * count_m - 2)].value != str(i):
                s_ws['H{}'.format(4 + 3 * count_m - 2)] = s_ws['H{}'.format(4 + 3 * count_m - 2)].value + '/' + str(i)
            else:
                s_ws['H{}'.format(4 + 3 * count_m - 2)] = str(i)

            ss = m.sample.filter(sample_date=i).distinct('sample_time').order_by('sample_time')  # 日期’i‘对应的样品表对象
            for s in ss:
                if str(s.sample_time) == '08:00:00':
                    s_ws['I{}'.format(4 + 3 * count_m - 2)] = '08:00:00'
                    s_ws['J{}'.format(4 + 3 * count_m - 2)] = s.sample_number
                    s_ws['K{}'.format(
                        4 + 3 * count_m - 2)] = s.sample_color + '、' + s.sample_odor + '、' + s.sample_turbidity
                    s_ws['L{}'.format(4 + 3 * count_m - 2)] = s.monitor_task
                    s_ws['M{}'.format(4 + 3 * count_m - 2)] = s.sample_count
                    s_ws['N{}'.format(4 + 3 * count_m - 2)] = s.people
                elif str(s.sample_time) == '12:30:00':
                    s_ws['I{}'.format(4 + 3 * count_m - 1)] = '12:30:00'
                    s_ws['J{}'.format(4 + 3 * count_m - 1)] = s.sample_number
                    s_ws['K{}'.format(
                        4 + 3 * count_m - 1)] = s.sample_color + '、' + s.sample_odor + '、' + s.sample_turbidity
                    s_ws['L{}'.format(4 + 3 * count_m - 1)] = s.monitor_task
                    s_ws['M{}'.format(4 + 3 * count_m - 1)] = s.sample_count
                    s_ws['N{}'.format(4 + 3 * count_m - 1)] = s.people
                else:
                    s_ws['I{}'.format(4 + 3 * count_m)] = '19:30:00'
                    s_ws['J{}'.format(4 + 3 * count_m)] = s.sample_number
                    s_ws['K{}'.format(
                        4 + 3 * count_m)] = s.sample_color + '、' + s.sample_odor + '、' + s.sample_turbidity
                    s_ws['L{}'.format(4 + 3 * count_m)] = s.monitor_task
                    s_ws['M{}'.format(4 + 3 * count_m)] = s.sample_count
                    s_ws['N{}'.format(4 + 3 * count_m)] = s.people
            fs = m.flow.filter(flow_date=i).distinct('flow_time').order_by('flow_time')  # 日期’i‘对应的流量表对象
            wb = openpyxl.load_workbook(os.path.join(my_setting.excel_folder, '流量表.xlsx'))
            ws = wb['流量']
            ws['A4'].value = m.name
            temp_ex_name = m.name + '_' + str(i) + '_' + '流量表.xlsx'
            monitor_flow_excel = os.path.join(output_path_per_monitor, temp_ex_name)
            day_avg_flow_lst = []
            for f in fs:  # 遍历某一采样时间的流量对象
                # print(f.flow_date, f.flow_time,f.volume1)
                if str(f.flow_time) == '08:00:00':
                    time_lst = [f.time1, f.time2, f.time3]
                    flow_lst = [f.volume1, f.volume2, f.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(4, 7):
                        temp_time = 'D' + str(j)
                        ws[temp_time].value = time_lst[j - 4]
                        temp_water = 'E' + str(j)
                        ws[temp_water].value = flow_lst[j - 4]
                elif str(f.flow_time) == '12:30:00':
                    time_lst = [f.time1, f.time2, f.time3]
                    flow_lst = [f.volume1, f.volume2, f.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(7, 10):
                        temp_time = 'D' + str(j)
                        ws[temp_time].value = time_lst[j - 7]
                        temp_water = 'E' + str(j)
                        ws[temp_water].value = flow_lst[j - 7]
                else:
                    time_lst = [f.time1, f.time2, f.time3]
                    flow_lst = [f.volume1, f.volume2, f.volume3]
                    flow1 = (flow_lst[0] / 1000) / time_lst[0]
                    flow2 = (flow_lst[1] / 1000) / time_lst[1]
                    flow3 = (flow_lst[2] / 1000) / time_lst[2]
                    day_avg_flow_lst.append((flow1 + flow2 + flow3) / 3)
                    for j in range(10, 13):
                        temp_time = 'D' + str(j)
                        ws[temp_time].value = time_lst[j - 10]
                        temp_water = 'E' + str(j)
                        ws[temp_water].value = flow_lst[j - 10]
            intro = '监测日期：{}                观测者：{}                检查者：'.format(i, m.people)
            ws['A2'] = intro
            wb.save(monitor_flow_excel)
            day_avg = mean(day_avg_flow_lst)
            day_total_flow = day_avg * 86400 / 1000
            monitor_flow_dic[str(i)] = day_total_flow
        avg_day_total_flow = mean(monitor_flow_dic.values())
        date_sorted_lst = sorted(monitor_flow_dic.keys())
        start_time = date_sorted_lst[0]
        end_time = date_sorted_lst[-1]
        f_ws['A{}'.format(3 + count_m)] = count_m
        f_ws['B{}'.format(3 + count_m)] = m.name
        f_ws['E{}'.format(3 + count_m)] = m.geophysical_point
        f_ws['L{}'.format(3 + count_m)] = start_time
        f_ws['M{}'.format(3 + count_m)] = end_time
        f_ws['N{}'.format(3 + count_m)] = '当日00:00至24:00'
        f_ws['O{}'.format(3 + count_m)] = avg_day_total_flow
        f_ws['Q{}'.format(3 + count_m)] = '容器法'
        f_ws['R{}'.format(3 + count_m)] = m.people

        s_ws['A{}'.format(4 + 3 * count_m - 2)] = count_m
        s_ws['B{}'.format(4 + 3 * count_m - 2)] = m.name
        s_ws['E{}'.format(4 + 3 * count_m - 2)] = m.geophysical_point
        m_wb.save(static_excel)
